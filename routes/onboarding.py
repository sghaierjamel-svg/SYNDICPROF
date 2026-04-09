import io, secrets
from flask import render_template, request, redirect, url_for, flash, send_file, jsonify
from core import app, db
from models import Organization, Block, Apartment, User
from utils import current_user, current_organization, login_required, admin_required, subscription_required


# ─── Dismiss du setup wizard ──────────────────────────────────────────────────

@app.route('/onboarding/dismiss', methods=['POST'])
@login_required
@admin_required
def onboarding_dismiss():
    org = current_organization()
    if org:
        org.setup_dismissed = True
        db.session.commit()
    return jsonify({'ok': True})


# ─── Téléchargement du modèle Excel ──────────────────────────────────────────

@app.route('/onboarding/template')
@login_required
@admin_required
def onboarding_template():
    """Génère et retourne un fichier Excel modèle à remplir."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Feuille 1 : Données à remplir ──────────────────────────────────────
    ws = wb.active
    ws.title = "Appartements & Residents"

    # Couleurs
    blue_fill   = PatternFill("solid", fgColor="1E40AF")
    green_fill  = PatternFill("solid", fgColor="065F46")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")
    gray_fill   = PatternFill("solid", fgColor="F3F4F6")
    white_font  = Font(color="FFFFFF", bold=True, size=11)
    header_border = Border(
        bottom=Side(style='medium', color='1E40AF'),
        right=Side(style='thin', color='CCCCCC'),
    )

    # En-têtes colonnes
    headers = [
        ("Bâtiment *",        "Nom du bâtiment (ex: A, Bloc 1, Tour Sud). Obligatoire.", 18),
        ("Appartement *",     "Numéro ou nom de l'appartement (ex: 101, RDC-G). Obligatoire.", 18),
        ("Charges/mois (DT)", "Montant des charges mensuelles en DT. Laisser vide = 100 DT.", 20),
        ("Parking",           "Numéro de place de parking (optionnel).", 14),
        ("Nom résident",      "Prénom et nom du résident (optionnel).", 22),
        ("Email résident",    "Email de connexion du résident (optionnel). Un compte sera créé.", 28),
        ("Téléphone WhatsApp","Numéro WhatsApp du résident pour les notifications (optionnel).", 22),
    ]

    for col_idx, (header, _, width) in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font  = white_font
        cell.fill  = blue_fill if col_idx <= 2 else (green_fill if col_idx <= 4 else PatternFill("solid", fgColor="374151"))
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = header_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 36

    # Ligne d'exemple (grisée)
    example = ["A", "101", 150, "P-01", "Ben Ali Mohamed", "benali@email.com", "21234567"]
    for col_idx, value in enumerate(example, start=1):
        cell = ws.cell(row=2, column=col_idx, value=value)
        cell.fill = yellow_fill
        cell.font = Font(italic=True, color="6B7280")
        cell.alignment = Alignment(horizontal='center')

    # Étiquette ligne exemple
    ws.cell(row=2, column=1).value = "A  ← EXEMPLE (supprimer cette ligne)"

    # Lignes vierges (30 lignes)
    for row in range(3, 33):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.fill = gray_fill if row % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            cell.alignment = Alignment(horizontal='center')

    # Figer la ligne d'en-tête
    ws.freeze_panes = "A2"

    # ── Feuille 2 : Instructions ───────────────────────────────────────────
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("GUIDE D'IMPORT SYNDICPRO", None),
        ("", None),
        ("1. COLONNES OBLIGATOIRES (en bleu)", None),
        ("   • Bâtiment  : nom du bâtiment (ex: A, B, Bloc 1). Sera créé automatiquement.", None),
        ("   • Appartement : numéro ou identifiant de l'appartement (ex: 101, RDC-G).", None),
        ("", None),
        ("2. COLONNES OPTIONNELLES", None),
        ("   • Charges/mois : montant en DT. Si vide, 100 DT par défaut.", None),
        ("   • Parking : numéro de place (ex: P-12). Peut être vide.", None),
        ("   • Nom résident : prénom + nom du résident.", None),
        ("   • Email résident : si renseigné, un compte résident sera créé.", None),
        ("     Un mot de passe temporaire sera généré et affiché après l'import.", None),
        ("   • Téléphone WhatsApp : format international (ex: 21612345678).", None),
        ("", None),
        ("3. RÈGLES IMPORTANTES", None),
        ("   • Supprimer la ligne d'exemple (ligne 2 jaune) avant d'importer.", None),
        ("   • Un même bâtiment peut apparaître sur plusieurs lignes.", None),
        ("   • Si un appartement existe déjà, il sera mis à jour (pas dupliqué).", None),
        ("   • Si un email résident existe déjà, le compte sera lié à l'appartement.", None),
        ("   • Les mots de passe temporaires sont affichés UNE SEULE FOIS après import.", None),
        ("     Notez-les ou transmettez-les aux résidents immédiatement.", None),
        ("", None),
        ("4. FORMAT DU FICHIER", None),
        ("   • Format accepté : .xlsx (Excel 2007 et supérieur).", None),
        ("   • Taille maximale : 5 Mo.", None),
        ("   • Ne pas modifier les en-têtes de colonnes.", None),
    ]
    ws2.column_dimensions['A'].width = 80
    for row_idx, (text, _) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if row_idx == 1:
            cell.font = Font(bold=True, size=14, color="1E40AF")
        elif text.startswith(("1.", "2.", "3.", "4.")):
            cell.font = Font(bold=True, size=11)
        else:
            cell.font = Font(size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='SyndicPro_Import_Modele.xlsx'
    )


# ─── Import Excel ─────────────────────────────────────────────────────────────

@app.route('/onboarding/import', methods=['GET', 'POST'])
@login_required
@admin_required
@subscription_required
def onboarding_import():
    org  = current_organization()
    user = current_user()

    if request.method == 'GET':
        return render_template('onboarding_import.html', user=user)

    # ── Validation fichier ────────────────────────────────────────────────
    f = request.files.get('excel_file')
    if not f or f.filename == '':
        flash('Veuillez sélectionner un fichier Excel.', 'warning')
        return redirect(url_for('onboarding_import'))

    if not f.filename.lower().endswith('.xlsx'):
        flash('Format non accepté. Utilisez un fichier .xlsx (Excel).', 'danger')
        return redirect(url_for('onboarding_import'))

    raw = f.read()
    if len(raw) > 5 * 1024 * 1024:
        flash('Fichier trop lourd (max 5 Mo).', 'danger')
        return redirect(url_for('onboarding_import'))

    # ── Lecture Excel ──────────────────────────────────────────────────────
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filename=io.BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        flash(f'Impossible de lire le fichier : {e}', 'danger')
        return redirect(url_for('onboarding_import'))

    # ── Traitement ligne par ligne ─────────────────────────────────────────
    results = {
        'blocks_created':     0,
        'apts_created':       0,
        'apts_updated':       0,
        'residents_created':  0,
        'residents_linked':   0,
        'errors':             [],
        'new_accounts':       [],   # [(nom, email, password_temp)]
    }

    # Charge les bâtiments existants en cache
    blocks_cache = {b.name.strip().upper(): b
                    for b in Block.query.filter_by(organization_id=org.id).all()}
    apts_cache   = {}  # (block_id, number_upper) → Apartment

    for apt in Apartment.query.filter_by(organization_id=org.id).all():
        apts_cache[(apt.block_id, apt.number.strip().upper())] = apt

    users_cache = {u.email.strip().lower(): u
                   for u in User.query.filter_by(organization_id=org.id).all()}

    row_num = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1

        # Ignorer lignes vides
        if not any(row):
            continue

        # Extraction colonnes
        block_name  = str(row[0]).strip() if row[0] is not None else ''
        apt_number  = str(row[1]).strip() if row[1] is not None else ''
        monthly_fee_raw = row[2]
        parking     = str(row[3]).strip() if row[3] is not None else None
        res_name    = str(row[4]).strip() if row[4] is not None else None
        res_email   = str(row[5]).strip().lower() if row[5] is not None else None
        res_phone   = str(row[6]).strip() if row[6] is not None else None

        # Ignorer ligne d'exemple
        if 'exemple' in block_name.lower() or 'example' in block_name.lower():
            continue

        # Validation obligatoire
        if not block_name or not apt_number:
            results['errors'].append(f"Ligne {row_num + 1} : Bâtiment et Appartement obligatoires.")
            continue

        # Charges
        try:
            monthly_fee = float(monthly_fee_raw) if monthly_fee_raw is not None else 100.0
            if monthly_fee <= 0:
                monthly_fee = 100.0
        except (ValueError, TypeError):
            monthly_fee = 100.0

        # Parking vide → None
        if parking in ('', 'None', 'nan'):
            parking = None

        # ── Bâtiment ──────────────────────────────────────────────────────
        block_key = block_name.upper()
        if block_key not in blocks_cache:
            block = Block(organization_id=org.id, name=block_name)
            db.session.add(block)
            db.session.flush()
            blocks_cache[block_key] = block
            results['blocks_created'] += 1
        else:
            block = blocks_cache[block_key]

        # ── Appartement ───────────────────────────────────────────────────
        apt_key = (block.id, apt_number.upper())
        if apt_key not in apts_cache:
            apt = Apartment(
                organization_id=org.id,
                block_id=block.id,
                number=apt_number,
                monthly_fee=monthly_fee,
                parking_spot=parking,
            )
            db.session.add(apt)
            db.session.flush()
            apts_cache[apt_key] = apt
            results['apts_created'] += 1
        else:
            apt = apts_cache[apt_key]
            apt.monthly_fee  = monthly_fee
            if parking:
                apt.parking_spot = parking
            results['apts_updated'] += 1

        # ── Résident ─────────────────────────────────────────────────────
        if res_email:
            if res_email in users_cache:
                # Compte existant → lier à l'appartement si pas déjà fait
                existing = users_cache[res_email]
                if existing.apartment_id != apt.id:
                    existing.apartment_id = apt.id
                    if res_phone and not existing.phone:
                        existing.phone = res_phone
                    results['residents_linked'] += 1
            else:
                # Nouveau compte résident
                pwd_temp = secrets.token_urlsafe(8)
                resident = User(
                    organization_id=org.id,
                    email=res_email,
                    name=res_name or res_email.split('@')[0],
                    role='resident',
                    apartment_id=apt.id,
                    phone=res_phone if res_phone and res_phone not in ('None', 'nan', '') else None,
                )
                resident.set_password(pwd_temp)
                db.session.add(resident)
                users_cache[res_email] = resident
                results['residents_created'] += 1
                results['new_accounts'].append({
                    'nom':   resident.name,
                    'email': res_email,
                    'pwd':   pwd_temp,
                    'apt':   f"{block_name}-{apt_number}",
                })

    try:
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de l\'enregistrement : {e}', 'danger')
        return redirect(url_for('onboarding_import'))

    return render_template('onboarding_import.html',
                           user=user,
                           results=results,
                           show_results=True)
