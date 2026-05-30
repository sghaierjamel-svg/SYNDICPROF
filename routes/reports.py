from flask import render_template, send_file, jsonify, request
from core import app, db
from models import Apartment, Payment, Expense, MiscReceipt
from utils import (current_user, current_organization, login_required,
                   subscription_required, last_n_months, get_month_name,
                   get_paid_months_map, get_unpaid_map, get_unpaid_details_map)
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from sqlalchemy import func
from sqlalchemy.orm import joinedload
import pandas as pd
import io


@app.route('/tresorerie')
@login_required
@subscription_required
def tresorerie():
    org = current_organization()
    months = last_n_months(12)
    apartments = (Apartment.query.options(joinedload(Apartment.block))
                  .filter_by(organization_id=org.id)
                  .order_by(Apartment.block_id, Apartment.number).all())

    # Fenêtre de la requête = premier mois affiché → on ne charge QUE les 12 mois,
    # pas les 10 ans d'historique. Agrégation SQL GROUP BY (pas de boucle Python).
    first_year, first_month = months[0]
    window_start = date(first_year, first_month, 1)

    def _mk(y, m):
        return f"{y}-{m:02d}"

    # Paiements agrégés par (appartement, année, mois)
    pay_rows = (db.session.query(
            Payment.apartment_id,
            func.extract('year',  Payment.payment_date).label('y'),
            func.extract('month', Payment.payment_date).label('m'),
            func.sum(Payment.amount).label('total'))
        .filter(Payment.organization_id == org.id,
                Payment.payment_date >= window_start)
        .group_by(Payment.apartment_id, 'y', 'm').all())
    pay_map = {(r.apartment_id, _mk(int(r.y), int(r.m))): float(r.total or 0) for r in pay_rows}

    # Dépenses agrégées par (année, mois)
    exp_rows = (db.session.query(
            func.extract('year',  Expense.expense_date).label('y'),
            func.extract('month', Expense.expense_date).label('m'),
            func.sum(Expense.amount).label('total'))
        .filter(Expense.organization_id == org.id,
                Expense.expense_date >= window_start)
        .group_by('y', 'm').all())
    exp_map = {_mk(int(r.y), int(r.m)): float(r.total or 0) for r in exp_rows}

    # Encaissements divers agrégés par (année, mois)
    misc_rows = (db.session.query(
            func.extract('year',  MiscReceipt.payment_date).label('y'),
            func.extract('month', MiscReceipt.payment_date).label('m'),
            func.sum(MiscReceipt.amount).label('total'))
        .filter(MiscReceipt.organization_id == org.id,
                MiscReceipt.payment_date >= window_start)
        .group_by('y', 'm').all())
    misc_map = {_mk(int(r.y), int(r.m)): float(r.total or 0) for r in misc_rows}

    data = []
    for apt in apartments:
        row = {'apartment': f"{apt.block.name}-{apt.number}", 'months': {}}
        for year, month in months:
            month_key = f"{year}-{month:02d}"
            row['months'][month_key] = pay_map.get((apt.id, month_key), 0)
        data.append(row)

    # Ligne encaissements divers
    misc_row = {'apartment': 'ENCAISSEMENTS DIVERS', 'months': {}}
    for year, month in months:
        month_key = f"{year}-{month:02d}"
        misc_row['months'][month_key] = misc_map.get(month_key, 0)

    expense_row = {'apartment': 'DÉPENSES', 'months': {}}
    for year, month in months:
        month_key = f"{year}-{month:02d}"
        expense_row['months'][month_key] = exp_map.get(month_key, 0)

    solde_row = {'apartment': 'SOLDE', 'months': {}}
    for year, month in months:
        month_key = f"{year}-{month:02d}"
        total_in = sum(row['months'][month_key] for row in data) + misc_row['months'][month_key]
        total_out = expense_row['months'][month_key]
        solde_row['months'][month_key] = total_in - total_out

    return render_template('tresorerie.html',
                         data=data,
                         misc_row=misc_row,
                         expense_row=expense_row,
                         solde_row=solde_row,
                         months=months,
                         user=current_user())


@app.route('/comptable')
@login_required
@subscription_required
def comptable():
    org = current_organization()
    today = date.today()

    # Années disponibles — DISTINCT en SQL (pas de chargement de tous les paiements)
    year_rows = (db.session.query(func.substr(Payment.month_paid, 1, 4))
                 .filter(Payment.organization_id == org.id)
                 .distinct().all())
    years_set = {today.year}
    for (ystr,) in year_rows:
        if ystr and len(ystr) >= 4:
            try:
                years_set.add(int(ystr))
            except (ValueError, TypeError):
                pass
    available_years = sorted(years_set, reverse=True)

    selected_year = request.args.get('year', '', type=str).strip()
    if selected_year and selected_year.isdigit() and int(selected_year) in years_set:
        year = int(selected_year)
        months = [(year, m) for m in range(1, 13)]
    else:
        selected_year = ''
        months = []
        for i in range(8, -1, -1):
            month_date = today - relativedelta(months=i)
            months.append((month_date.year, month_date.month))
        for i in range(1, 4):
            month_date = today + relativedelta(months=i)
            months.append((month_date.year, month_date.month))

    apartments = (Apartment.query.options(joinedload(Apartment.block))
                  .filter_by(organization_id=org.id)
                  .order_by(Apartment.block_id, Apartment.number).all())

    # 1 seule requête pour les mois payés, réutilisée pour la grille ET les impayés
    all_paid_months = get_paid_months_map(org.id)
    unpaid_map = get_unpaid_map(org.id, apartments, paid=all_paid_months)
    data = []

    for apt in apartments:
        row = {
            'apartment': f"{apt.block.name}-{apt.number}",
            'monthly_fee': apt.monthly_fee,
            'credit_balance': apt.credit_balance,
            'months': {}
        }

        apt_paid_months = all_paid_months.get(apt.id, set())

        for year, month in months:
            month_key = f"{year}-{month:02d}"
            paid = month_key in apt_paid_months
            amount = apt.monthly_fee if paid else 0
            row['months'][month_key] = {'paid': paid, 'amount': amount}

        row['unpaid_count'] = unpaid_map.get(apt.id, 0)
        data.append(row)

    return render_template('comptable.html', data=data, months=months, user=current_user(),
                           available_years=available_years, selected_year=selected_year)


@app.route('/export_excel')
@login_required
@subscription_required
def export_excel():
    from openpyxl.styles import PatternFill, Font
    from models import User as UserModel

    org = current_organization()

    # Années disponibles — DISTINCT en SQL (paiements via month_paid + dépenses via année)
    years_set = {str(date.today().year)}
    for (ystr,) in (db.session.query(func.substr(Payment.month_paid, 1, 4))
                    .filter(Payment.organization_id == org.id).distinct().all()):
        if ystr and len(ystr) >= 4:
            years_set.add(ystr)
    for (yr,) in (db.session.query(func.extract('year', Expense.expense_date))
                  .filter(Expense.organization_id == org.id).distinct().all()):
        if yr:
            years_set.add(str(int(yr)))
    available_years = sorted(years_set, reverse=True)

    year_param = request.args.get('year', '')
    if year_param not in available_years:
        return render_template('export_excel_choose.html',
                               available_years=available_years,
                               current_year=str(date.today().year),
                               user=current_user())

    year_str = year_param
    year = int(year_str)

    # Données de l'année sélectionnée — requêtes scopées (pas de chargement des 10 ans)
    # month_paid pilote l'année (les paiements en avance tombent dans la bonne année)
    payments_year = (Payment.query
                     .options(joinedload(Payment.apartment).joinedload(Apartment.block))
                     .filter(Payment.organization_id == org.id,
                             Payment.month_paid.like(f"{year_str}-%"))
                     .order_by(Payment.payment_date.desc(), Payment.id.desc()).all())
    expenses_year = (Expense.query
                     .filter(Expense.organization_id == org.id,
                             func.extract('year', Expense.expense_date) == year)
                     .order_by(Expense.expense_date.desc()).all())
    apartments = (Apartment.query.options(joinedload(Apartment.block))
                  .filter_by(organization_id=org.id)
                  .order_by(Apartment.block_id, Apartment.number).all())

    # Impayés de TOUS les appartements en 1 requête (au lieu de 7 appels N+1 par apt)
    unpaid_details = get_unpaid_details_map(org.id, apartments)

    # ---- Sheet 1: Appartements (état actuel) ----
    residents_by_apt = {u.apartment_id: u for u in UserModel.query.filter_by(organization_id=org.id, role='resident').all()}
    df_apartments = pd.DataFrame([{
        'Bloc': apt.block.name,
        'Appartement': apt.number,
        'Référence': f"{apt.block.name}-{apt.number}",
        'Place Parking': apt.parking_spot or '',
        'Redevance Mensuelle (DT)': apt.monthly_fee,
        'Crédit (DT)': apt.credit_balance,
        'Résident': residents_by_apt[apt.id].name if apt.id in residents_by_apt else '',
        'Email Résident': residents_by_apt[apt.id].email if apt.id in residents_by_apt else '',
        'Téléphone': residents_by_apt[apt.id].phone if apt.id in residents_by_apt else '',
        'Mois Impayés': unpaid_details.get(apt.id, (0, ''))[0],
        'Total Dû (DT)': apt.monthly_fee * unpaid_details.get(apt.id, (0, ''))[0],
        'Créé le': apt.created_at.strftime('%d/%m/%Y') if apt.created_at else '',
    } for apt in apartments])

    # ---- Sheet 2: Encaissements YYYY ----
    MODE_LABELS = {'especes': 'Espèces', 'virement': 'Virement', 'cheque': 'Chèque'}
    df_payments = pd.DataFrame([{
        'ID': p.id,
        'Appartement': f"{p.apartment.block.name}-{p.apartment.number}" if p.apartment else '',
        'Montant (DT)': p.amount,
        'Date Paiement': p.payment_date.strftime('%d/%m/%Y'),
        'Mois Payé': p.month_paid,
        'Mode': MODE_LABELS.get(p.payment_mode or 'especes', 'Espèces'),
        'N° Chèque': p.cheque_number or '',
        'Banque': p.cheque_bank or '',
        'Crédit Utilisé': p.credit_used or 0,
        'Description': p.description or '',
    } for p in payments_year])
    if df_payments.empty:
        df_payments = pd.DataFrame(columns=['ID', 'Appartement', 'Montant (DT)', 'Date Paiement', 'Mois Payé', 'Mode', 'N° Chèque', 'Banque', 'Crédit Utilisé', 'Description'])

    # ---- Sheet 3: Dépenses YYYY ----
    df_expenses = pd.DataFrame([{
        'ID': e.id,
        'Montant (DT)': e.amount,
        'Date': e.expense_date.strftime('%d/%m/%Y'),
        'Catégorie': e.category or '',
        'Description': e.description or '',
    } for e in expenses_year])
    if df_expenses.empty:
        df_expenses = pd.DataFrame(columns=['ID', 'Montant (DT)', 'Date', 'Catégorie', 'Description'])

    # ---- Sheet 4: Impayés (état actuel) ----
    df_unpaid = pd.DataFrame([{
        'Appartement': f"{apt.block.name}-{apt.number}",
        'Place Parking': apt.parking_spot or '',
        'Redevance Mensuelle (DT)': apt.monthly_fee,
        'Crédit Disponible (DT)': apt.credit_balance,
        'Mois Impayés': unpaid_details.get(apt.id, (0, ''))[0],
        'Prochain Mois': unpaid_details.get(apt.id, (0, ''))[1],
        'Total Dû (DT)': apt.monthly_fee * unpaid_details.get(apt.id, (0, ''))[0],
    } for apt in apartments if unpaid_details.get(apt.id, (0, ''))[0] > 0])
    if df_unpaid.empty:
        df_unpaid = pd.DataFrame(columns=['Appartement', 'Place Parking', 'Redevance Mensuelle (DT)', 'Crédit Disponible (DT)', 'Mois Impayés', 'Prochain Mois', 'Total Dû (DT)'])

    # ---- Sheet 5: Tableau Comptable YYYY (12 mois fixes, avec couleurs) ----
    MONTH_NAMES = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Juin', 'Juil', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']

    # Paid months per apartment for the selected year (payments_year déjà scopé à l'année)
    paid_by_apt = {}
    for p in payments_year:
        if p.month_paid:
            paid_by_apt.setdefault(p.apartment_id, set()).add(p.month_paid)

    comptable_data = []
    for apt in apartments:
        apt_paid = paid_by_apt.get(apt.id, set())
        nb_payes = 0
        total_percu = 0.0
        row = {
            'Appartement': f"{apt.block.name}-{apt.number}",
            'Redevance (DT)': apt.monthly_fee,
        }
        for m in range(1, 13):
            month_key = f"{year_str}-{m:02d}"
            if month_key in apt_paid:
                row[MONTH_NAMES[m - 1]] = 'Payé'
                nb_payes += 1
                total_percu += apt.monthly_fee
            else:
                row[MONTH_NAMES[m - 1]] = 'Impayé'
        nb_impayes = 12 - nb_payes
        row['Nb Payés'] = nb_payes
        row['Nb Impayés'] = nb_impayes
        row['Total Perçu (DT)'] = round(total_percu, 3)
        row['Total Dû (DT)'] = round(apt.monthly_fee * nb_impayes, 3)
        comptable_data.append(row)
    df_comptable = pd.DataFrame(comptable_data)

    # Write Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_apartments.to_excel(writer, sheet_name='Appartements', index=False)
        df_payments.to_excel(writer, sheet_name=f'Encaissements {year_str}', index=False)
        df_expenses.to_excel(writer, sheet_name=f'Dépenses {year_str}', index=False)
        df_unpaid.to_excel(writer, sheet_name='Impayés', index=False)
        df_comptable.to_excel(writer, sheet_name=f'Tableau {year_str}', index=False)

        # Color coding for Tableau sheet
        fill_green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        font_green = Font(color='276221')
        fill_red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        font_red = Font(color='9C0006')

        ws_tab = writer.sheets[f'Tableau {year_str}']
        headers = [cell.value for cell in ws_tab[1]]
        month_cols = {h: i + 1 for i, h in enumerate(headers) if h in MONTH_NAMES}

        for row_idx in range(2, ws_tab.max_row + 1):
            for col_idx in month_cols.values():
                cell = ws_tab.cell(row=row_idx, column=col_idx)
                if cell.value == 'Payé':
                    cell.fill = fill_green
                    cell.font = font_green
                elif cell.value == 'Impayé':
                    cell.fill = fill_red
                    cell.font = font_red

        ws_tab.freeze_panes = 'C2'

        # Auto-fit column widths on all sheets
        for ws in writer.sheets.values():
            for col in ws.columns:
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 32)

    output.seek(0)
    filename = f"SyndicPro_{org.name}_{year_str}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, download_name=filename, as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
